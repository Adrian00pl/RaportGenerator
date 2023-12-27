from PyQt6.QtWidgets import (QApplication, QMainWindow, QPushButton, QLabel,QHBoxLayout, QVBoxLayout, QWidget,
                             QToolBar,QStackedWidget, QRadioButton,QFrame,QSplitter,QTableWidget,QTableWidgetItem,QFileDialog,QComboBox,QCheckBox, QDateEdit)
from PyQt6.QtCore import Qt, QSize
import sys

class FirstWindow(QWidget):
    """
    This "window" is a QWidget. If it has no parent, it
    will appear as a free-floating window as we want.
    """

    def load_excel_data(self, table):
        file_name, _ = QFileDialog.getOpenFileName(self, "Select Excel file", "", "Excel Files (*.xlsx)")

        if file_name:
            from openpyxl import load_workbook

            wb = load_workbook(file_name)
            sheet = wb.active

            table.setRowCount(sheet.max_row-1)
            table.setColumnCount(sheet.max_column)

            # Set headers for rows (excluding the first row)
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value is not None:
                    table.setHorizontalHeaderItem(col - 1, QTableWidgetItem(str(cell_value)))

            # Set headers for columns (excluding the first column)
            # for row in range(1, sheet.max_row + 1):
            #     cell_value = sheet.cell(row=row, column=1).value
            #     if cell_value is not None:
            #         table.setVerticalHeaderItem(row - 1, QTableWidgetItem(str(cell_value)))

            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        table.setItem(row - 2, col - 1, QTableWidgetItem(str(cell_value)))

    def count_values(self, table, column_index):
        
    def sort_table(self, index):
        self.stacked_widget.currentWidget().sortItems(index)

    def change_table(self, index):
        self.stacked_widget.setCurrentIndex(index)

    def __init__(self):
        super().__init__()
        layout0 = QHBoxLayout()
        layout1 = QVBoxLayout()
        layout2 = QVBoxLayout()

        date_label1 = QLabel("Select date range:")
        date_from1 = QDateEdit()
        date_to1 = QDateEdit()

        date_label2 = QLabel("Select date range:")
        date_from2 = QDateEdit()
        date_to2 = QDateEdit()

        date_label3 = QLabel("Select date range:")
        date_from3 = QDateEdit()
        date_to3 = QDateEdit()

        self.button1 = QPushButton('Income', self)
        self.button2 = QPushButton('Cost', self)
        self.button3 = QPushButton('Result', self)
        self.button4 = QPushButton('Generuj', self)
        self.button5 = QPushButton('Pobierz dane', self)
        self.button6 = QPushButton('Pobierz dane', self)

        self.button1.clicked.connect(lambda: self.change_table(0))
        self.button2.clicked.connect(lambda: self.change_table(1))
        self.button3.clicked.connect(lambda: self.change_table(2))
        self.button4.clicked.connect(self.count_values)
        self.button5.clicked.connect(lambda: self.load_excel_data(self.table1))
        self.button6.clicked.connect(lambda: self.load_excel_data(self.table2))

        table_layout1 = QVBoxLayout()
        self.table1 = QTableWidget()
        self.table11 = QTableWidget()
        self.table1.horizontalHeader().sectionDoubleClicked.connect(lambda index: self.count_values(self.table1, index))
        table_layout1.addWidget(self.table1)
        table_layout1.addWidget(date_label1)
        table_layout1.addWidget(date_from1)
        table_layout1.addWidget(date_to1)
        table_layout1.addWidget(QLabel("Select an option:"))
        table_layout1.addWidget(QCheckBox("Option 1"))
        table_layout1.addWidget(QCheckBox("Option 2"))
        table_layout1.addWidget(QCheckBox("Option 3"))
        table_layout1.addWidget(self.button5)
        table_layout1.addWidget(self.table11)

        table_layout2 = QVBoxLayout()
        self.table2 = QTableWidget()
        table_layout2.addWidget(self.table2)
        table_layout2.addWidget(date_label2)
        table_layout2.addWidget(date_from2)
        table_layout2.addWidget(date_to2)
        table_layout2.addWidget(self.button6)

        table_layout3 = QVBoxLayout()
        self.table3 = QTableWidget()
        table_layout3.addWidget(self.table3)
        table_layout3.addWidget(date_label3)
        table_layout3.addWidget(date_from3)
        table_layout3.addWidget(date_to3)
        table_layout3.addWidget(self.button4)

        self.stacked_widget = QStackedWidget()
        self.stacked_widget.addWidget(QWidget())  # Pusty widget - miejsce na indeks 0
        self.stacked_widget.addWidget(QWidget())  # Pusty widget - miejsce na indeks 1
        self.stacked_widget.addWidget(QWidget())  # Pusty widget - miejsce na indeks 2

        self.stacked_widget.widget(0).setLayout(table_layout1)
        self.stacked_widget.widget(1).setLayout(table_layout2)
        self.stacked_widget.widget(2).setLayout(table_layout3)

        for i in range(5):
            for j in range(5):
                item = QTableWidgetItem(f"Row {i} Col {j}")
                self.table1.setItem(i, j, item)

        self.sort_label = QLabel("Sort by:")
        self.sort_combobox = QComboBox()
        self.sort_combobox.addItems(["Column 1", "Column 2", "Column 3", "Column 4", "Column 5"])
        self.sort_combobox.currentIndexChanged.connect(self.sort_table)

        layout2.addWidget(self.sort_label)
        layout2.addWidget(self.sort_combobox)

        layout1.addWidget(self.button1)
        layout1.addWidget(self.button2)
        layout1.addWidget(self.button3)
        layout1.addStretch()

        layout2.addWidget(self.stacked_widget)
        layout2.addStretch()

        layout0.addLayout(layout1, stretch=1)
        layout0.addLayout(layout2, stretch=4)

        self.setLayout(layout0)

